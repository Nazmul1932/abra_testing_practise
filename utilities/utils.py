import inspect
import logging
import softest
import csv
from openpyxl import Workbook, load_workbook


class Utils(softest.TestCase):
    def custom_logger(logLevel=logging.INFO):
        logger_name = inspect.stack()[1][3]
        logger = logging.getLogger(logger_name)
        logger.setLevel(logLevel)
        fh = logging.FileHandler("/media/user/New Volume/abra_testing_practise/reports/automation.log", mode='a')
        formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(name)s : %(message)s',
                                      datefmt='%m/%d/%Y %I:%M:%S %p')
        fh.setFormatter(formatter)
        logger.addHandler(fh)
        return logger

    def read_data_from_excel(file_name, sheet):
        datalist = []
        wb = load_workbook(filename=file_name)
        sh = wb[sheet]
        row_ct = sh.max_row
        col_ct = sh.max_column

        for i in range(2, row_ct + 1):
            row = []
            for j in range(1, col_ct + 1):
                row.append(sh.cell(row=i, column=j).value)
            datalist.append(row)
        return datalist

    def read_data_from_csv(filename):
        # Create an empty list
        datalist = []
        # Open CSV file
        csvdata = open(filename, "r")
        # Create CSV reader
        reader = csv.reader(csvdata)
        # skip header
        next(reader)
        # Add CSV rows to list
        for rows in reader:
            datalist.append(rows)
        return datalist


